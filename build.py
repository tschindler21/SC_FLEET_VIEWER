#!/usr/bin/env python3
BUILD_VERSION = "2026-05-16-v10"
"""
SC Fleet Viewer — Build Script
Reads fleet-template.xlsx + ships.json + images/ and generates index.html

Usage: python build.py
"""

import json
import base64
import os
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


import csv
import glob

try:
    from PIL import Image
    import io
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("WARNING: Pillow not installed. Images will be embedded at full size.")
    print("         Install with: pip install Pillow")

# ── Configuration ──
EXCEL_FILE = "fleet-template.xlsx"
SHIPS_FILE = "ships.json"
IMAGES_DIR = "images"
OUTPUT_FILE = "index.html"
IMAGE_WIDTH = 560
IMAGE_QUALITY = 72

# ── Manufacturer accent colors ──
MFR_COLORS = {
    "RSI": "#4fc3f7", "Drake": "#ffd740", "Aegis": "#ff5252",
    "Origin": "#b0bec5", "Crusader": "#69f0ae", "Anvil": "#ff9800",
    "MISC": "#ce93d8", "Consolidated Outland": "#ffab40",
    "Kruger": "#e0e0e0", "Argo": "#80cbc4",
}
ROLE_COLORS = {
    "Combat": "#ff5252", "Hauling": "#69f0ae", "Industrial": "#ffd740",
    "Exploration": "#4fc3f7", "Multi-Role": "#b0bec5", "Support": "#ce93d8",
    "Mil. Support": "#ff9800",
}
SIZE_ORDER = {"Small": 0, "Medium": 1, "Large": 2, "Capital": 3}


def load_ships_db():
    with open(SHIPS_FILE) as f:
        data = json.load(f)
    return data["ships"]


def load_excel():
    wb = load_workbook(EXCEL_FILE)

    # Read org info
    ws_org = wb["Org"]
    org = {
        "name": ws_org["B3"].value or "My Org",
        "motto": ws_org["B4"].value or "",
        "description": ws_org["B5"].value or "",
        "logo": ws_org["B6"].value or "",
    }

    # Read fleet
    ws_fleet = wb["Fleet"]
    members = {}
    for row in ws_fleet.iter_rows(min_row=2, max_col=6, values_only=True):
        vals = list(row) + [None] * (6 - len(row))  # pad if fewer columns
        member, ship_name, melt_val, store_val, notes, img_file = vals
        if not member or not ship_name:
            continue
        member = str(member).strip()
        ship_name = str(ship_name).strip()
        if member not in members:
            members[member] = []
        members[member].append({
            "ship_name": ship_name,
            "melt_value": int(melt_val) if melt_val else 0,
            "store_value": int(store_val) if store_val else 0,
            "notes": str(notes).strip() if notes else "",
            "img_file": str(img_file).strip() if img_file else "",
        })

    return org, members


def load_org_json():
    """Load org info from org.json (simpler alternative to Excel)."""
    if os.path.exists("org.json"):
        with open("org.json") as f:
            return json.load(f)
    return None


def _lookup_store_price(item_name, ships_db):
    """Look up current store price from ships.json by item name.
    Uses aliases and fuzzy matching. Returns price as int, or 0 if not found."""
    item_clean = item_name.strip()
    
    # Check aliases first (loaded from ships.json _aliases)
    sid = _resolve_ship_id(item_clean, ships_db)
    if sid and sid in ships_db:
        return ships_db[sid].get("store_price", 0)
    
    return 0


def _resolve_ship_id(item_name, ships_db):
    """Resolve a ship name (from CSV export) to a ships.json ID."""
    item_clean = item_name.strip()
    item_lower = item_clean.lower()
    
    # 1. Direct alias lookup
    # Load aliases from the db module-level (passed alongside ships)
    global _aliases_cache
    if not hasattr(_resolve_ship_id, '_aliases'):
        try:
            with open(SHIPS_FILE) as f:
                import json as _json
                full_db = _json.load(f)
                _resolve_ship_id._aliases = full_db.get("_aliases", {})
        except:
            _resolve_ship_id._aliases = {}
    
    aliases = _resolve_ship_id._aliases
    if item_clean in aliases:
        return aliases[item_clean]
    
    # 2. Exact match on ship name
    for sid, ship in ships_db.items():
        if ship["name"].lower() == item_lower:
            return sid
    
    # 3. Match without manufacturer prefix
    for sid, ship in ships_db.items():
        db_name = ship["name"].lower()
        db_short = db_name.split(" ", 1)[-1] if " " in db_name else db_name
        if item_lower == db_short or item_lower in db_name:
            return sid
    
    # 4. Loose keyword match
    item_words = set(item_lower.replace("-", " ").split())
    item_words -= {"mk", "ii", "mk2", "l", "22"}  # noise words
    for sid, ship in ships_db.items():
        db_words = set(ship["name"].lower().replace("-", " ").split())
        db_words -= {"mk", "ii", "mk2"}
        if item_words and item_words.issubset(db_words):
            return sid
    
    return None


def load_hangars_csv(ships_db):
    """Load member fleets from CSV files in hangars/ folder.
    
    Smart pipeline: groups raw CSV rows by pledge, extracts the primary items
    (ships, CCUs), ignores noise (hangars, flair, insurance items), and 
    calculates warbond savings.
    
    Filename becomes member name (e.g., hangars/Matt_Graver.csv → "Matt Graver")
    Files starting with SAMPLE_ are skipped.
    """
    hangars_dir = "hangars"
    if not os.path.exists(hangars_dir):
        return None
    
    csv_files = glob.glob(os.path.join(hangars_dir, "*.csv"))
    csv_files = [f for f in csv_files if not os.path.basename(f).startswith("SAMPLE_")]
    
    if not csv_files:
        return None
    
    # Item types we care about as primary fleet items
    PRIMARY_TYPES = {"ship", "ccu"}
    # Items to always ignore (noise)
    NOISE_KEYWORDS = {"self-land hangar", "aeroview hangar", "revel & york hangar",
                      "vfg industrial hangar", "poster", "model ", "trophy", "pennant",
                      "manual", "digital download", "name reservation", "best in show",
                      "starting money", "star citizen digital", "squadron 42 digital"}
    
    members = {}
    
    for csv_file in csv_files:
        basename = os.path.splitext(os.path.basename(csv_file))[0]
        member_name = basename.replace("_", " ")
        print(f"  Loading {csv_file} → {member_name}")
        
        # Phase 1: Read all rows
        raw_rows = []
        with open(csv_file, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                continue
            
            # Auto-detect columns
            col_map = {}
            for h in reader.fieldnames:
                hl = h.strip().lower()
                if hl in ("item name", "name", "contains", "ship name", "ship", "title"):
                    col_map["name"] = h
                elif hl in ("pledge name", "pledge", "pledge type"):
                    col_map["pledge"] = h
                elif hl in ("melt value", "value", "melt", "credit value"):
                    col_map["melt"] = h
                elif hl in ("store value", "current value", "pledge value", "price"):
                    col_map["store"] = h
                elif hl in ("type", "item type", "category", "kind"):
                    col_map["type"] = h
                elif hl in ("insurance", "ins"):
                    col_map["insurance"] = h
                elif hl in ("manufacturer", "mfr"):
                    col_map["mfr"] = h
                elif hl in ("date", "pledge date", "acquired"):
                    col_map["date"] = h
                elif hl in ("member name", "member", "owner", "player"):
                    col_map["member"] = h
            
            # If member column exists, use it; otherwise use filename
            has_member_col = "member" in col_map
            # If no type column, assume all rows are ships (simplified org export)
            has_type_col = "type" in col_map
            
            for row in reader:
                # Skip separator rows from fleet planner export
                first_val = list(row.values())[0] if row else ""
                if first_val.strip() in ("---", "STORE PRICES", "STORE"):
                    continue
                
                if has_member_col:
                    row_member = row.get(col_map["member"], "").strip()
                    if row_member:
                        member_name = row_member
                
                pledge = row.get(col_map.get("pledge", ""), "").strip()
                name = row.get(col_map.get("name", ""), "").strip()
                melt_str = row.get(col_map.get("melt", ""), "0").strip()
                store_str = row.get(col_map.get("store", ""), "0").strip()
                item_type = row.get(col_map.get("type", ""), "").strip().lower()
                insurance = row.get(col_map.get("insurance", ""), "").strip()
                mfr = row.get(col_map.get("mfr", ""), "").strip()
                date = row.get(col_map.get("date", ""), "").strip()
                
                # If no type column, default to "ship"
                if not has_type_col and name:
                    item_type = "ship"
                
                melt = float(melt_str.replace("$", "").replace(",", "").strip() or "0")
                store = float(store_str.replace("$", "").replace(",", "").strip() or "0")
                
                # If no price data, look up from ships.json
                if melt == 0 and store == 0 and name:
                    looked_up = _lookup_store_price(name, ships_db)
                    if looked_up:
                        store = looked_up
                        melt = looked_up
                
                if name:
                    raw_rows.append({
                        "pledge": pledge or name, "name": name, "type": item_type,
                        "melt": melt, "store": store, "insurance": insurance,
                        "mfr": mfr, "date": date,
                    })
        
        # Phase 2: Group by pledge
        pledges = {}
        for row in raw_rows:
            key = row["pledge"]
            if key not in pledges:
                pledges[key] = {"rows": [], "melt": row["melt"], "store": row["store"],
                                "date": row["date"]}
            pledges[key]["rows"].append(row)
        
        # Phase 3: Extract clean fleet items per pledge
        fleet_items = []
        
        for pledge_name, pledge_data in pledges.items():
            rows = pledge_data["rows"]
            melt = pledge_data["melt"]
            store = pledge_data["store"]
            date = pledge_data["date"]
            
            # Find insurance from any row in this pledge
            insurance = ""
            for r in rows:
                if r["insurance"]:
                    insurance = r["insurance"]
                    break
            
            # Find primary items (ships and CCUs)
            primary_items = []
            for r in rows:
                # Skip noise items
                name_lower = r["name"].lower()
                if any(noise in name_lower for noise in NOISE_KEYWORDS):
                    continue
                if r["type"] in PRIMARY_TYPES:
                    primary_items.append(r)
            
            # Also check for paints (they're valuable and meltable)
            paint_items = [r for r in rows if r["type"] == "paint"]
            
            if primary_items:
                # Pledge has ships or CCUs — output those
                for item in primary_items:
                    # Look up store value from ships database
                    store = _lookup_store_price(item["name"], ships_db) or int(melt)
                    saved = int(store - melt) if store > melt else 0
                    fleet_items.append({
                        "ship_name": item["name"],
                        "pledge_name": pledge_name,
                        "melt_value": int(melt),
                        "store_value": store,
                        "saved": saved,
                        "insurance": insurance,
                        "mfr": item["mfr"],
                        "date": date,
                        "item_type": item["type"],
                        "img_file": "",
                    })
            elif paint_items and melt > 0:
                # Paint-only pledge (standalone paint purchase)
                for item in paint_items:
                    fleet_items.append({
                        "ship_name": item["name"],
                        "pledge_name": pledge_name,
                        "melt_value": int(melt),
                        "store_value": int(store),
                        "saved": 0,
                        "insurance": "",
                        "mfr": item["mfr"],
                        "date": date,
                        "item_type": "paint",
                        "img_file": "",
                    })
            elif melt > 0:
                # Meltable pledge with no ships/CCUs/paints (gear pack, etc.)
                # Summarize as one item
                fleet_items.append({
                    "ship_name": pledge_name,
                    "pledge_name": pledge_name,
                    "melt_value": int(melt),
                    "store_value": int(store),
                    "saved": 0,
                    "insurance": "",
                    "mfr": "",
                    "date": date,
                    "item_type": "gear",
                    "img_file": "",
                })
            # else: $0 melt pledge (free flair, event rewards) — skip entirely
        
        members[member_name] = fleet_items
        
        # Print summary
        ships = [i for i in fleet_items if i["item_type"] == "ship"]
        ccus = [i for i in fleet_items if i["item_type"] == "ccu"]
        paints = [i for i in fleet_items if i["item_type"] == "paint"]
        gear = [i for i in fleet_items if i["item_type"] == "gear"]
        total_melt = sum(i["melt_value"] for i in fleet_items)
        print(f"    {len(raw_rows)} raw rows → {len(fleet_items)} items "
              f"({len(ships)} ships, {len(ccus)} CCUs, {len(paints)} paints, {len(gear)} gear packs)")
        print(f"    Total melt value: ${total_melt}")
    
    return members if members else None


def find_ship_id(ship_name, ships_db):
    """Find ship ID by name, using aliases and fuzzy matching."""
    resolved = _resolve_ship_id(ship_name, ships_db)
    if resolved:
        return resolved
    # Last resort: partial match
    for sid, ship in ships_db.items():
        if ship_name.lower() in ship["name"].lower():
            return sid
    return None


def load_image_b64(filepath):
    """Load an image, resize, and return as base64 data URI."""
    if not os.path.exists(filepath):
        return None
    try:
        if HAS_PIL:
            img = Image.open(filepath).convert("RGB")
            w, h = img.size
            new_h = int(h * IMAGE_WIDTH / w)
            img = img.resize((IMAGE_WIDTH, new_h), Image.LANCZOS)
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=IMAGE_QUALITY)
            b64 = base64.b64encode(buf.getvalue()).decode()
        else:
            with open(filepath, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()
        return f"data:image/jpeg;base64,{b64}"
    except Exception as e:
        print(f"  WARNING: Could not load image {filepath}: {e}")
        return None


def collect_images(members, ships_db):
    """Collect all ship images from the images/ directory."""
    images = {}

    # Try to load org logo
    for ext in [".jpg", ".jpeg", ".png", ".webp"]:
        logo_path = os.path.join(IMAGES_DIR, f"logo{ext}")
        if os.path.exists(logo_path):
            img = load_image_b64(logo_path)
            if img:
                images["logo"] = img
                print(f"  Logo: {logo_path}")
            break

    # Collect ship images
    seen_ships = set()
    for member, ships in members.items():
        for entry in ships:
            sid = find_ship_id(entry["ship_name"], ships_db)
            if not sid or sid in seen_ships:
                continue
            seen_ships.add(sid)

            # Priority: explicit img_file > ship_id named file > skip
            candidates = []
            if entry["img_file"]:
                candidates.append(os.path.join(IMAGES_DIR, entry["img_file"]))
            for ext in [".jpg", ".jpeg", ".png", ".webp"]:
                candidates.append(os.path.join(IMAGES_DIR, f"{sid}{ext}"))

            for path in candidates:
                if os.path.exists(path):
                    img = load_image_b64(path)
                    if img:
                        images[sid] = img
                        print(f"  {sid}: {path}")
                    break

    return images


def generate_card(ship, accent, images, owner=None, delay_idx=0):
    """Generate HTML for a single ship card."""
    sid = ship.get("_id", "")
    has_img = sid in images
    img_html = (
        f'<div class="hero"><img src="{images[sid]}" alt="{ship["name"]}"><div class="hero-fade"></div></div>'
        if has_img else '<div class="no-img"></div>'
    )
    info_class = "info has-img" if has_img else "info"
    owner_html = f'<div class="ship-owner">Owner: <span>{owner}</span></div>' if owner else ""

    highlights = []
    for k, v in ship.get("details", {}).items():
        if k not in ("Note", "Announced", "Availability", "Series"):
            highlights.append(f"{k}: {v}")

    hl_html = '<div class="highlights">'
    for h in highlights[:4]:
        hl_html += f'<span class="hl" style="color:{accent};background:{accent}0a;border:1px solid {accent}20;">{h}</span>'
    hl_html += '</div>'

    det_html = f'<div class="details-toggle" style="color:{accent};" onclick="this.classList.toggle(\'open\');this.nextElementSibling.classList.toggle(\'open\');"><span class="arrow">▶</span> DETAILED SPECS</div><div class="details-panel"><div class="details-grid">'
    for dk, dv in ship.get("details", {}).items():
        full = " full" if len(str(dv)) > 35 else ""
        vc = "#556677" if dk == "Note" else "#7a8a9a"
        det_html += f'<div class="detail-item{full}"><span class="detail-key">{dk}</span><span class="detail-val" style="color:{vc};">{dv}</span></div>'
    det_html += '</div></div>'

    return f'''<div class="card" style="border:1px solid {accent}20;border-left:2px solid {accent};animation-delay:{delay_idx * 0.06}s;"><div class="accent-bar" style="background:linear-gradient(90deg,{accent},{accent}44,transparent);"></div>{img_html}<div class="{info_class}"><div class="tags"><span class="mfr-tag" style="color:{accent};background:{accent}15;">{ship["mfr"].upper()}</span><span class="role-tag">{ship.get("subrole", ship["role"]).upper()}</span><span class="size-tag">{ship["size"].upper()}</span></div><div class="ship-name">{ship["name"]}</div>{owner_html}<div class="ship-desc">{ship.get("desc", "")}</div>{hl_html}{det_html}</div></div>
'''


def build_html(org, members, ships_db, images):
    """Generate the complete HTML file."""

    # Build flat list of all ships
    all_ships = []
    for member_name, ships in members.items():
        for entry in ships:
            sid = find_ship_id(entry["ship_name"], ships_db)
            if sid:
                ship = dict(ships_db[sid])
                ship["_id"] = sid
                all_ships.append({"ship": ship, "owner": member_name})
            else:
                print(f"  WARNING: Ship '{entry['ship_name']}' not found in database, skipping")

    total_ships = len(all_ships)
    total_members = len(members)
    member_list = list(members.keys())
    member_ids = [m.replace(" ", "_").lower() for m in member_list]
    short_names = [m.split(" ")[0][:5].upper() for m in member_list]

    # Count sizes
    large = sum(1 for s in all_ships if s["ship"]["size"] in ("Large", "Capital"))
    medium = sum(1 for s in all_ships if s["ship"]["size"] == "Medium")
    small = sum(1 for s in all_ships if s["ship"]["size"] == "Small")

    logo_html = f'<img class="org-logo" src="{images["logo"]}" alt="{org["name"]}">' if "logo" in images else ""

    # ── Start building HTML ──
    css = open(os.path.join(os.path.dirname(__file__), "style.css")).read() if os.path.exists("style.css") else DEFAULT_CSS

    h = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{org["name"]} — Fleet Viewer</title>
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>{DEFAULT_CSS}</style>
</head>
<body>
<div class="glow"></div>
<div class="wrap">
<div class="org-header">{logo_html}
<div class="org-name">{org["name"]}</div>
<div class="org-motto">{org["motto"]}</div>
<div class="org-desc">{org["description"]}</div>
<div class="org-line"></div></div>
<div class="summary">
<div class="summary-item"><div class="s-val">{total_members}</div><div class="s-lbl">MEMBERS</div></div>
<div class="summary-item"><div class="s-val">{total_ships}</div><div class="s-lbl">SHIPS</div></div>
<div class="summary-item"><div class="s-val">{large}</div><div class="s-lbl">LARGE</div></div>
<div class="summary-item"><div class="s-val">{medium}</div><div class="s-lbl">MEDIUM</div></div>
<div class="summary-item"><div class="s-val">{small}</div><div class="s-lbl">SMALL</div></div>
</div>
'''

    # Fleet Overview button (prominent)
    h += '<div style="text-align:center;margin-bottom:16px;"><a href="fleet-overview.html" style="display:inline-block;padding:10px 28px;font-family:\'DM Mono\',monospace;font-size:11px;letter-spacing:3px;border:1px solid rgba(0,224,255,0.3);border-radius:6px;background:rgba(0,224,255,0.05);color:#00e0ff;text-decoration:none;transition:all 0.3s;" onmouseover="this.style.background=\'rgba(0,224,255,0.12)\';this.style.borderColor=\'#00e0ff\';" onmouseout="this.style.background=\'rgba(0,224,255,0.05)\';this.style.borderColor=\'rgba(0,224,255,0.3)\';">&#x1F680; FLEET OVERVIEW</a></div>'

    # Tabs
    h += '<div class="tabs"><div class="tab active" onclick="switchTab(\'org\')">ORG FLEET</div>'
    for m_name in member_list:
        mid = m_name.replace(" ", "_").lower()
        h += f'<div class="tab" onclick="switchTab(\'{mid}\')">{m_name.upper()}</div>'
    h += '<div class="tab" onclick="switchTab(\'compare\')">⚖ COMPARE</div></div>'

    # ── Org Fleet Tab ──
    h += '<div class="tab-content active" id="tab-org">'
    h += '<div class="sort-bar"><span class="sort-label">SORT BY</span>'
    h += '<span class="sort-btn active" onclick="sortOrg(\'role\')">ROLE</span>'
    h += '<span class="sort-btn" onclick="sortOrg(\'size\')">SIZE</span>'
    h += '<span class="sort-btn" onclick="sortOrg(\'manufacturer\')">MANUFACTURER</span></div>'

    # Member filter buttons
    h += '<div class="member-filters"><span class="sort-label">MEMBERS</span>'
    for m_name, short in zip(member_list, short_names):
        h += f'<span class="mbr-btn active" onclick="toggleMember(this,\'{m_name}\')">{short}</span>'
    h += '</div>'

    # Sort views
    for sort_type in ["role", "size", "manufacturer"]:
        display = "block" if sort_type == "role" else "none"
        h += f'<div class="sort-view" id="sort-{sort_type}" style="display:{display};">'

        groups = {}
        for entry in all_ships:
            ship = entry["ship"]
            if sort_type == "role":
                key = ship["role"]
            elif sort_type == "size":
                key = ship["size"]
            else:
                key = ship["mfr"]
            groups.setdefault(key, []).append(entry)

        if sort_type == "size":
            sorted_keys = sorted(groups.keys(), key=lambda x: SIZE_ORDER.get(x, 99), reverse=True)
        else:
            sorted_keys = sorted(groups.keys())

        idx = 0
        for key in sorted_keys:
            entries = sorted(groups[key], key=lambda e: e["ship"]["name"])
            color = ROLE_COLORS.get(key, "#c89564") if sort_type == "role" else "#c89564"
            h += f'<div class="cat-header" style="color:{color};">{key.upper()} — {len(entries)}</div><div class="cards">'
            for entry in entries:
                accent = MFR_COLORS.get(entry["ship"]["mfr"], "#c89564")
                h += generate_card(entry["ship"], accent, images, owner=entry["owner"], delay_idx=idx)
                idx += 1
            h += '</div>'
        h += '</div>'
    h += '</div>'

    # ── Member Tabs ──
    for m_name in member_list:
        mid = m_name.replace(" ", "_").lower()
        initials = "".join(w[0] for w in m_name.split()[:2]).upper()
        bc = list(MFR_COLORS.values())[hash(m_name) % len(MFR_COLORS)]
        member_ships = [e for e in all_ships if e["owner"] == m_name]

        h += f'<div class="tab-content" id="tab-{mid}">'
        h += f'<div class="member-header"><div class="member-avatar" style="background:{bc}22;color:{bc};border:1px solid {bc}44;">{initials}</div>'
        h += f'<div class="member-info"><div class="member-name">{m_name}</div>'
        h += f'<div class="member-count">{len(member_ships)} SHIP{"S" if len(member_ships) != 1 else ""}</div></div></div>'
        h += '<div class="cards">'
        for idx, entry in enumerate(member_ships):
            accent = MFR_COLORS.get(entry["ship"]["mfr"], "#c89564")
            h += generate_card(entry["ship"], accent, images, delay_idx=idx)
        h += '</div></div>'

    # ── Compare Tab ──
    h += '<div class="tab-content" id="tab-compare">'
    h += '<div style="margin-bottom:14px;"><div style="font-size:14px;font-weight:700;color:#eef1f5;margin-bottom:10px;">Fleet Comparison</div>'
    h += '<div style="font-size:9px;color:#667;font-family:\'DM Mono\',monospace;margin-bottom:12px;">Select members to compare fleets side by side.</div>'
    h += '<div class="member-filters" id="compare-filters"><span class="sort-label">SELECT</span>'
    for m_name, short in zip(member_list, short_names):
        h += f'<span class="mbr-btn" onclick="toggleCompare(this,\'{m_name}\')">{short}</span>'
    h += '</div></div><div id="compare-table-wrap"><div style="text-align:center;padding:40px 0;color:#445;font-family:\'DM Mono\',monospace;font-size:10px;">Select members above to compare fleets</div></div></div>'

    # ── Footer ──
    h += f'<div class="footer">{org["name"]} — {org["motto"]}<br><span style="color:#334455;">BUILD {BUILD_VERSION} | UNOFFICIAL FAN PROJECT — NOT AFFILIATED WITH CIG</span><br><a href="https://buymeacoffee.com/schindi21" target="_blank" style="color:#889;text-decoration:none;border-bottom:1px solid #556;">buy me a coffee</a> &bull; BTC: <span style="color:#778;font-size:10px;">bc1qdyh5g2zska7s9e4vu27hzqyyre60t6khl4srnx</span></div></div>'

    # ── JavaScript ──
    member_ships_js = "{"
    for m_name in member_list:
        ships_list = ",".join([f'"{e["ship"]["name"]}"' for e in all_ships if e["owner"] == m_name])
        member_ships_js += f'"{m_name}":[{ships_list}],'
    member_ships_js += "}"

    ship_details_js = "{"
    seen = set()
    for entry in all_ships:
        ship = entry["ship"]
        if ship["name"] in seen:
            continue
        seen.add(ship["name"])
        details = {k: v for k, v in ship.get("details", {}).items() if k not in ("Note", "Announced", "Series")}
        ship_details_js += f'"{ship["name"]}":{json.dumps({"size": ship["size"], "role": ship["role"], **details})},'
    ship_details_js += "}"

    mi = str(member_ids)

    h += f'''
<script>
function switchTab(id){{document.querySelectorAll(".tab-content").forEach(t=>t.classList.remove("active"));document.querySelectorAll(".tab").forEach(t=>t.classList.remove("active"));const safeId=id==="org"?"tab-org":id==="compare"?"tab-compare":"tab-"+id;document.getElementById(safeId).classList.add("active");const names=["org"].concat({mi}).concat(["compare"]);const idx=names.indexOf(id);if(idx>=0)document.querySelectorAll(".tab")[idx].classList.add("active");document.getElementById(safeId).querySelectorAll(".card").forEach((c,i)=>{{c.style.animation="none";c.offsetHeight;c.style.animation="slideIn 0.4s ease "+(i*0.06)+"s both";}});}}
function sortOrg(type){{document.querySelectorAll(".sort-view").forEach(v=>v.style.display="none");document.getElementById("sort-"+type).style.display="block";document.querySelectorAll(".sort-btn").forEach(b=>b.classList.remove("active"));event.target.classList.add("active");applyMemberFilter();}}
const hiddenMembers=new Set();function toggleMember(btn,name){{if(hiddenMembers.has(name)){{hiddenMembers.delete(name);btn.classList.add("active");btn.style.background="rgba(200,149,100,0.1)";btn.style.borderColor="rgba(200,149,100,0.4)";btn.style.color="#c89564";}}else{{hiddenMembers.add(name);btn.classList.remove("active");btn.style.background="rgba(255,255,255,0.02)";btn.style.borderColor="rgba(255,255,255,0.06)";btn.style.color="#667";}}applyMemberFilter();}}
function applyMemberFilter(){{document.querySelectorAll("#tab-org .card").forEach(c=>{{const o=c.querySelector(".ship-owner span");if(!o)return;c.style.display=hiddenMembers.has(o.textContent)?"none":"";}});document.querySelectorAll("#tab-org .cat-header").forEach(h=>{{const cards=h.nextElementSibling;if(!cards||!cards.classList.contains("cards"))return;const vis=Array.from(cards.children).filter(c=>c.style.display!=="none").length;h.innerHTML=h.innerHTML.replace(/\\u2014\\s*\\d+/,"\\u2014 "+vis);h.style.display=vis===0?"none":"";cards.style.display=vis===0?"none":"";}});let i=0;document.querySelectorAll("#tab-org .sort-view[style*=block] .card").forEach(c=>{{if(c.style.display!=="none"){{c.style.animation="none";c.offsetHeight;c.style.animation="slideIn 0.4s ease "+(i*0.06)+"s both";i++;}}}});}}
const memberShips={member_ships_js};const shipInfo={ship_details_js};const compareSet=new Set();
function toggleCompare(btn,name){{if(compareSet.has(name)){{compareSet.delete(name);btn.classList.remove("active");btn.style.background="rgba(255,255,255,0.02)";btn.style.borderColor="rgba(255,255,255,0.06)";btn.style.color="#667";}}else{{compareSet.add(name);btn.classList.add("active");btn.style.background="rgba(200,149,100,0.1)";btn.style.borderColor="rgba(200,149,100,0.4)";btn.style.color="#c89564";}}renderCompare();}}
function renderCompare(){{const wrap=document.getElementById("compare-table-wrap");if(compareSet.size===0){{wrap.innerHTML='<div style="text-align:center;padding:40px 0;color:#445;font-family:DM Mono,monospace;font-size:10px;">Select members above to compare fleets</div>';return;}}const sel=Array.from(compareSet);const as=new Set();sel.forEach(m=>memberShips[m]?.forEach(s=>as.add(s)));const sizeOrd={{Large:0,Medium:1,Small:2,Capital:0}};const ships=Array.from(as).sort((a,b)=>{{const sa=shipInfo[a]||{{}},sb=shipInfo[b]||{{}};if((sizeOrd[sa.size]||2)!==(sizeOrd[sb.size]||2))return(sizeOrd[sa.size]||2)-(sizeOrd[sb.size]||2);return a.localeCompare(b);}});const detailKeys=new Set();ships.forEach(s=>{{const info=shipInfo[s]||{{}};Object.keys(info).forEach(k=>{{if(k!=="size"&&k!=="role")detailKeys.add(k);}});}});const cols=Array.from(detailKeys);const colColors={{"Shields":"#ce93d8","Cargo":"#ffd740","Weapons":"#ff5252","Missiles":"#ff9800","Torpedoes":"#ff5252","Quantum Fuel":"#4fc3f7","QFuel":"#4fc3f7","SCM Speed":"#69f0ae","Turrets":"#ff9800","Medical":"#69f0ae","Mining":"#ffd740","Salvage":"#ffd740"}};let t='<div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;font-family:DM Mono,monospace;font-size:9px;"><tr><th style="text-align:left;padding:8px 6px;border-bottom:2px solid rgba(200,149,100,0.2);color:#c89564;font-size:10px;">SHIP</th><th style="padding:8px 4px;border-bottom:2px solid rgba(200,149,100,0.2);color:#667;font-size:8px;">SIZE</th>';cols.forEach(c=>{{t+='<th style="padding:8px 4px;border-bottom:2px solid rgba(200,149,100,0.2);color:'+(colColors[c]||"#889")+';font-size:7px;letter-spacing:0.5px;white-space:nowrap;">'+c.toUpperCase()+'</th>';}});sel.forEach(m=>{{t+='<th style="padding:8px 4px;border-bottom:2px solid rgba(200,149,100,0.2);color:#c89564;font-size:8px;min-width:36px;">'+m.split(" ")[0].substring(0,5).toUpperCase()+'</th>';}});t+='</tr>';ships.forEach((ship,i)=>{{const info=shipInfo[ship]||{{}};const bg=i%2===0?'rgba(255,255,255,0.01)':'transparent';const sc={{Large:'#ff9800',Medium:'#4fc3f7',Small:'#69f0ae',Capital:'#ff5252'}}[info.size]||'#667';const shortName=ship.split(" ").slice(1).join(" ");t+='<tr style="background:'+bg+';"><td style="padding:6px;border-bottom:1px solid rgba(255,255,255,0.03);color:#dde;font-size:10px;white-space:nowrap;">'+shortName+'</td><td style="padding:6px 4px;border-bottom:1px solid rgba(255,255,255,0.03);color:'+sc+';text-align:center;font-size:8px;">'+(info.size||"")+'</td>';cols.forEach(c=>{{const val=info[c]||"\\u2014";const dim=val==="\\u2014";t+='<td style="padding:6px 4px;border-bottom:1px solid rgba(255,255,255,0.03);color:'+(dim?"#333":"#8899aa")+';text-align:center;font-size:8px;white-space:nowrap;">'+val+'</td>';}});sel.forEach(m=>{{const has=memberShips[m]?.includes(ship);t+='<td style="padding:6px 4px;border-bottom:1px solid rgba(255,255,255,0.03);text-align:center;">'+(has?'<span style="color:#69f0ae;">\\u25cf</span>':'<span style="color:#222;">\\u2014</span>')+'</td>';}});t+='</tr>';}});t+='<tr style="border-top:2px solid rgba(200,149,100,0.2);"><td style="padding:8px 6px;color:#c89564;font-weight:700;">TOTALS</td><td colspan="'+cols.length+'"></td>';sel.forEach(m=>{{t+='<td style="padding:8px 4px;text-align:center;color:#c89564;font-weight:700;">'+(memberShips[m]?.length||0)+'</td>';}});t+='</tr></table></div>';wrap.innerHTML=t;}}
</script>
</body></html>'''

    return h
DEFAULT_CSS = '''*{margin:0;padding:0;box-sizing:border-box;}body{background:#060a10;color:#e8edf2;font-family:"Space Grotesk",sans-serif;min-height:100vh;}body::before{content:"";position:fixed;inset:0;pointer-events:none;opacity:0.03;background-image:linear-gradient(rgba(0,224,255,0.3) 1px,transparent 1px),linear-gradient(90deg,rgba(0,224,255,0.3) 1px,transparent 1px);background-size:60px 60px;z-index:0;}.glow{position:fixed;top:-100px;left:50%;transform:translateX(-50%);width:500px;height:300px;background:radial-gradient(ellipse,rgba(0,224,255,0.06) 0%,transparent 70%);pointer-events:none;}.wrap{position:relative;z-index:10;max-width:580px;margin:0 auto;padding:24px 16px;}.org-header{text-align:center;margin-bottom:20px;animation:fadeIn 0.6s ease;}.org-logo{width:90px;height:90px;border-radius:50%;border:2px solid #c8956430;margin:0 auto 12px;display:block;}.org-name{font-size:36px;font-weight:700;letter-spacing:4px;color:#fff;}.org-motto{font-size:9px;letter-spacing:4px;color:#c89564;font-family:"DM Mono",monospace;margin-top:4px;opacity:0.7;}.org-desc{font-size:10px;color:#667788;font-family:"DM Mono",monospace;margin-top:8px;line-height:1.5;max-width:400px;margin-left:auto;margin-right:auto;}.org-line{width:60px;height:2px;background:linear-gradient(90deg,#c89564,#c8956440);margin:14px auto 0;}.summary{display:flex;justify-content:space-between;padding:10px 16px;margin-bottom:16px;background:rgba(200,149,100,0.03);border:1px solid rgba(200,149,100,0.08);border-radius:6px;}.summary-item{text-align:center;}.s-val{font-size:16px;font-weight:700;color:#c89564;line-height:1.2;}.s-lbl{font-size:7px;color:#445566;font-family:"DM Mono",monospace;letter-spacing:2px;margin-top:2px;}.tabs{display:flex;gap:4px;margin-bottom:16px;flex-wrap:wrap;}.tab{padding:6px 12px;font-size:9px;font-family:"DM Mono",monospace;letter-spacing:1px;border:1px solid rgba(255,255,255,0.06);border-radius:4px;background:rgba(255,255,255,0.02);color:#667;cursor:pointer;white-space:nowrap;transition:all 0.2s;user-select:none;}.tab:hover{border-color:rgba(200,149,100,0.3);color:#aab;}.tab.active{background:rgba(200,149,100,0.1);border-color:rgba(200,149,100,0.4);color:#c89564;}.sort-bar{display:flex;gap:6px;margin-bottom:14px;align-items:center;flex-wrap:wrap;}.sort-label{font-size:8px;color:#556;font-family:"DM Mono",monospace;letter-spacing:1px;}.sort-btn,.mbr-btn{padding:4px 10px;font-size:8px;font-family:"DM Mono",monospace;letter-spacing:1px;border:1px solid rgba(255,255,255,0.06);border-radius:3px;background:rgba(255,255,255,0.02);color:#667;cursor:pointer;transition:all 0.2s;user-select:none;}.sort-btn:hover,.mbr-btn:hover{border-color:rgba(200,149,100,0.3);color:#aab;}.sort-btn.active,.mbr-btn.active{background:rgba(200,149,100,0.1);border-color:rgba(200,149,100,0.4);color:#c89564;}.cat-header{font-size:11px;font-family:"DM Mono",monospace;letter-spacing:2px;padding:8px 0 6px;margin-top:12px;border-bottom:1px solid rgba(255,255,255,0.04);margin-bottom:8px;}.tab-content{display:none;}.tab-content.active{display:block;}.cards{display:flex;flex-direction:column;gap:10px;}.card{background:rgba(8,12,20,0.85);backdrop-filter:blur(12px);border-radius:8px;overflow:hidden;position:relative;animation:slideIn 0.4s ease both;}.accent-bar{height:3px;}.hero{position:relative;width:100%;height:160px;overflow:hidden;}.hero img{width:100%;height:100%;object-fit:cover;display:block;}.hero-fade{position:absolute;inset:0;background:linear-gradient(180deg,transparent 30%,rgba(8,12,20,0.95) 100%);}.no-img{height:6px;}.info{padding:0 16px 14px;position:relative;z-index:1;}.info.has-img{margin-top:-36px;}.tags{display:flex;align-items:center;gap:6px;margin-bottom:3px;flex-wrap:wrap;}.mfr-tag{font-size:8px;font-weight:700;font-family:"DM Mono",monospace;padding:1px 5px;border-radius:3px;letter-spacing:1px;}.role-tag{font-size:7px;color:#556677;font-family:"DM Mono",monospace;letter-spacing:2px;}.size-tag{font-size:7px;font-family:"DM Mono",monospace;padding:1px 4px;border-radius:2px;letter-spacing:1px;color:#556;background:rgba(255,255,255,0.04);}.ship-name{font-size:18px;font-weight:700;color:#eef1f5;letter-spacing:0.5px;line-height:1.15;margin-bottom:2px;}.ship-desc{font-size:10px;color:#7a8a9a;font-family:"DM Mono",monospace;line-height:1.5;margin:6px 0 10px;}.ship-owner{font-size:9px;color:#556677;font-family:"DM Mono",monospace;margin-bottom:4px;}.ship-owner span{color:#8899aa;}.highlights{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:12px;}.hl{font-size:8px;font-family:"DM Mono",monospace;letter-spacing:1px;padding:3px 8px;border-radius:3px;}.details-toggle{display:flex;align-items:center;gap:6px;margin-top:10px;padding:8px 0 0;border-top:1px solid rgba(255,255,255,0.04);cursor:pointer;user-select:none;font-size:9px;font-family:"DM Mono",monospace;letter-spacing:1px;}.details-toggle:hover{filter:brightness(1.3);}.details-toggle .arrow{display:inline-block;transition:transform 0.3s ease;font-size:10px;}.details-toggle.open .arrow{transform:rotate(90deg);}.details-panel{max-height:0;overflow:hidden;transition:max-height 0.4s cubic-bezier(0.16,1,0.3,1);}.details-panel.open{max-height:800px;}.details-grid{padding:10px 0 4px;display:grid;grid-template-columns:1fr 1fr;gap:4px 16px;}.detail-item{display:flex;flex-direction:column;padding:4px 0;border-bottom:1px solid rgba(255,255,255,0.02);}.detail-item.full{grid-column:1/-1;}.detail-key{font-size:7px;color:#445566;font-family:"DM Mono",monospace;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:1px;}.detail-val{font-size:10px;font-family:"DM Mono",monospace;}.member-header{display:flex;align-items:center;gap:12px;margin-bottom:14px;padding:10px 0;border-bottom:1px solid rgba(255,255,255,0.04);}.member-avatar{width:36px;height:36px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:700;}.member-info{flex:1;}.member-name{font-size:18px;font-weight:700;color:#eef1f5;}.member-count{font-size:9px;color:#556677;font-family:"DM Mono",monospace;letter-spacing:1px;}.member-filters{display:flex;flex-wrap:wrap;gap:4px;margin-bottom:14px;align-items:center;}.footer{text-align:center;margin-top:28px;opacity:0.25;font-size:8px;letter-spacing:3px;color:#556677;font-family:"DM Mono",monospace;}@keyframes fadeIn{from{opacity:0;}to{opacity:1;}}@keyframes slideIn{from{opacity:0;transform:translateY(12px);}to{opacity:1;transform:translateY(0);}}'''


def build_planner(members, ships_db):
    """Generate standalone planner.html with CCU planning tool."""
    
    # Build data structures for JS
    planner_data = {}
    for m_name, items in members.items():
        planner_data[m_name] = []
        for entry in items:
            planner_data[m_name].append({
                "name": entry["ship_name"],
                "pledge": entry.get("pledge_name", ""),
                "melt": entry.get("melt_value", 0),
                "store": entry.get("store_value", 0),
                "type": entry.get("item_type", "ship"),
                "insurance": entry.get("insurance", ""),
            })
    
    # Ship price lookup for CCU dropdowns
    ship_prices = {}
    for sid, ship in ships_db.items():
        p = ship.get("store_price", 0)
        if p > 0:
            ship_prices[ship["name"]] = p
    
    all_ships_sorted = sorted(
        [{"n": s["name"], "m": s["mfr"], "p": s.get("store_price", 0)} for s in ships_db.values()],
        key=lambda x: x["n"]
    )
    
    member_names = list(members.keys())
    short_names = [m.split(" ")[0][:5].upper() for m in member_names]
    
    css = '''*{margin:0;padding:0;box-sizing:border-box;}body{background:#060a10;color:#e8edf2;font-family:"Space Grotesk",sans-serif;min-height:100vh;}
.wrap{max-width:580px;margin:0 auto;padding:24px 16px;}
h1{font-size:28px;font-weight:700;letter-spacing:2px;color:#fff;text-align:center;margin-bottom:4px;}
.sub{font-size:9px;letter-spacing:5px;color:#c89564;font-family:"DM Mono",monospace;text-align:center;opacity:0.6;margin-bottom:24px;}
.summary{background:rgba(200,149,100,0.04);border:1px solid rgba(200,149,100,0.12);border-radius:8px;padding:14px;display:flex;justify-content:space-between;flex-wrap:wrap;gap:10px;margin-bottom:16px;}
.s-item{text-align:center;flex:1 1 80px;}.s-val{font-size:20px;font-weight:700;line-height:1.2;}.s-lbl{font-size:7px;color:#556;font-family:"DM Mono",monospace;letter-spacing:2px;margin-top:2px;}
.sec-head{font-size:11px;font-weight:700;color:#c89564;font-family:"DM Mono",monospace;letter-spacing:2px;border-bottom:1px solid rgba(200,149,100,0.15);padding-bottom:6px;margin-bottom:8px;display:flex;justify-content:space-between;}
.card{background:rgba(8,12,20,0.85);border:1px solid rgba(255,255,255,0.06);border-radius:6px;padding:8px 12px;margin-bottom:5px;display:flex;align-items:center;gap:8px;}
.mbr-btn{padding:4px 10px;font-size:8px;font-family:"DM Mono",monospace;letter-spacing:1px;border:1px solid rgba(255,255,255,0.06);border-radius:3px;background:rgba(255,255,255,0.02);color:#667;cursor:pointer;transition:all 0.2s;}
.mbr-btn:hover{border-color:rgba(200,149,100,0.3);color:#aab;}.mbr-btn.active{background:rgba(200,149,100,0.1);border-color:rgba(200,149,100,0.4);color:#c89564;}
.tag{font-size:8px;font-family:"DM Mono",monospace;padding:2px 6px;border-radius:3px;letter-spacing:1px;}
select{background:#0d1117;border:1px solid rgba(255,255,255,0.15);border-radius:4px;padding:6px 8px;color:#eef1f5;font-family:"DM Mono",monospace;font-size:11px;outline:none;flex:1;max-width:200px;}
select option{background:#0d1117;color:#eef1f5;}
input[type=number]{background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.1);border-radius:4px;padding:6px 10px;color:#eef1f5;font-family:"DM Mono",monospace;font-size:12px;outline:none;}
input[type=checkbox]{width:16px;height:16px;accent-color:#ff5252;cursor:pointer;}
.btn{padding:5px 12px;font-size:9px;font-family:"DM Mono",monospace;letter-spacing:1px;border:1px solid rgba(200,149,100,0.3);border-radius:4px;background:rgba(200,149,100,0.1);color:#c89564;cursor:pointer;}
.btn-x{padding:3px 8px;font-size:9px;font-family:"DM Mono",monospace;border:1px solid rgba(255,82,82,0.3);border-radius:4px;background:rgba(255,82,82,0.08);color:#ff5252;cursor:pointer;}
.footer{text-align:center;margin-top:28px;opacity:0.25;font-size:8px;letter-spacing:3px;color:#556677;font-family:"DM Mono",monospace;}'''

    h = f'''<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>CCU Fleet Planner</title>
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>{css}</style></head><body>
<div class="wrap">
<h1>CCU FLEET PLANNER</h1>
<div class="sub">PRIVATE \u2014 NOT SHOWN ON ORG PAGE</div>
<div style="display:flex;flex-wrap:wrap;gap:4px;margin-bottom:16px;align-items:center;">
<span style="font-size:8px;color:#556;font-family:DM Mono,monospace;letter-spacing:1px;">MEMBER</span>'''
    
    for m_name, short in zip(member_names, short_names):
        h += f'<span class="mbr-btn" onclick="selMbr(\'{m_name}\',this)">{short}</span>'
    
    h += '</div><div id="planner-summary"></div><div id="planner-area"><div style="text-align:center;padding:40px 0;color:#445;font-family:DM Mono,monospace;font-size:10px;">Select a member above to start planning</div></div>'
    h += '<div class="footer">CCU FLEET PLANNER \u2014 PRIVATE TOOL \u2014 NOT AFFILIATED WITH CIG</div></div>'

    h += f'''
<script>
const PF={json.dumps(planner_data)};
const SP={json.dumps(ship_prices)};
const ASL={json.dumps(all_ships_sorted)};
let PM=null,PMelts=new Set(),PCredit=0,PCcus=[];

function selMbr(n,btn){{document.querySelectorAll(".mbr-btn").forEach(b=>{{b.classList.remove("active");}});btn.classList.add("active");PM=n;PMelts=new Set();PCredit=0;PCcus=[];render();}}

function addCcu(){{const f=document.getElementById("cf").value,t=document.getElementById("ct").value;if(!f||!t)return;PCcus.push({{from:f,to:t,fromP:SP[f]||0,toP:SP[t]||0}});render();}}

function render(){{
const area=document.getElementById("planner-area"),sum=document.getElementById("planner-summary");
if(!PM||!PF[PM])return;
const fleet=PF[PM],meltT=fleet.filter((_,i)=>PMelts.has(i)).reduce((s,f)=>s+f.melt,0);
const totC=meltT+PCredit,ccuT=PCcus.reduce((s,c)=>s+Math.max(0,(c.toP||0)-(c.fromP||0)),0),rem=totC-ccuT,ok=rem>=0;

sum.innerHTML='<div class="summary"><div class="s-item"><div class="s-val" style="color:#69f0ae;">$'+totC+'</div><div class="s-lbl">TOTAL CREDIT</div></div><div class="s-item"><div class="s-val" style="color:#ff5252;">$'+ccuT+'</div><div class="s-lbl">CCU COST</div></div><div class="s-item"><div class="s-val" style="color:'+(ok?"#4fc3f7":"#ff5252")+';">$'+rem+'</div><div class="s-lbl">'+(ok?"REMAINING":"SHORT")+'</div></div></div>';

let h='<div class="sec-head"><span>CURRENT FLEET \u2014 '+PM+'</span></div>';
fleet.forEach((ship,i)=>{{const m=PMelts.has(i);
h+='<div class="card" style="border-left:2px solid '+(m?"#ff5252":"rgba(255,255,255,0.06)")+';opacity:'+(m?0.5:1)+';"><input type=checkbox '+(m?"checked":"")+" onchange=\\"PMelts.has("+i+")?PMelts.delete("+i+"):PMelts.add("+i+");render();\\" ><div style=\\"flex:1;\\"><div style=\\"font-size:13px;font-weight:600;color:#eef1f5;\\">"+(m?'<span class=\\"tag\\" style=\\"color:#ff5252;background:rgba(255,82,82,0.1);border:1px solid rgba(255,82,82,0.2);margin-right:6px;\\">MELT</span>':"")+ship.name+"</div>";
if(ship.insurance)h+='<div style="font-size:8px;color:#556;font-family:DM Mono,monospace;">'+ship.insurance+'</div>';
h+="</div><div style=\\"display:flex;gap:12px;\\"><div style=\\"font-size:11px;font-family:DM Mono,monospace;\\"><span style=\\"color:#889;\\">melt </span><span style=\\"color:"+(m?"#ff5252":"#69f0ae")+";\\">$"+ship.melt+"</span></div><div style=\\"font-size:11px;font-family:DM Mono,monospace;\\"><span style=\\"color:#889;\\">value </span><span style=\\"color:#4fc3f7;\\">$"+ship.store+"</span></div></div></div>";}});

h+='<div class="sec-head" style="margin-top:16px;"><span>STORE CREDIT</span></div><div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;"><span style="font-size:11px;color:#556;font-family:DM Mono,monospace;">$</span><input type=number value="'+PCredit+'" onchange="PCredit=Number(this.value)||0;render();" style="width:100px;"><span style="font-size:9px;color:#667;font-family:DM Mono,monospace;">+ $'+meltT+' from melts = <span style="color:#69f0ae;">$'+totC+'</span></span></div>';

h+='<div class="sec-head" style="margin-top:16px;"><span>CCU PLAN</span><span style="font-size:9px;color:#667;font-weight:400;">'+PCcus.length+' upgrades</span></div>';
PCcus.forEach((c,i)=>{{const cost=Math.max(0,(c.toP||0)-(c.fromP||0)),inv=c.toP>0&&c.fromP>0&&c.toP<=c.fromP;
h+='<div class="card" style="border-left:2px solid '+(inv?"#ff5252":"#4fc3f7")+';flex-wrap:wrap;"><div style="display:flex;align-items:center;gap:8px;flex:1;flex-wrap:wrap;"><span style="font-size:12px;font-weight:600;color:#eef1f5;">'+c.from+'</span><span style="color:#556;">\u2192</span><span style="font-size:12px;font-weight:600;color:#4fc3f7;">'+c.to+'</span></div><span style="font-size:14px;font-weight:700;font-family:DM Mono,monospace;color:'+(inv?"#ff5252":"#ffd740")+';">'+(inv?"INVALID":"$"+cost)+'</span><button class="btn-x" onclick="PCcus.splice('+i+',1);render();">\u2715</button><div style="display:flex;gap:12px;margin-top:8px;width:100%;align-items:center;"><span style="font-size:8px;color:#556;font-family:DM Mono,monospace;">FROM $</span><input type=number value="'+c.fromP+'" onchange="PCcus['+i+'].fromP=Number(this.value)||0;render();" style="width:70px;"><span style="font-size:8px;color:#556;font-family:DM Mono,monospace;">TO $</span><input type=number value="'+c.toP+'" onchange="PCcus['+i+'].toP=Number(this.value)||0;render();" style="width:70px;"></div></div>';}});

h+='<div style="display:flex;gap:6px;align-items:center;margin-top:8px;flex-wrap:wrap;"><select id="cf"><option value="">From...</option>';
ASL.forEach(s=>{{if(s.p>0)h+='<option value="'+s.n+'">'+s.n+' ($'+s.p+')</option>';else h+='<option value="'+s.n+'">'+s.n+'</option>';}});
h+='</select><span style="color:#556;">\u2192</span><select id="ct"><option value="">To...</option>';
ASL.forEach(s=>{{if(s.p>0)h+='<option value="'+s.n+'">'+s.n+' ($'+s.p+')</option>';else h+='<option value="'+s.n+'">'+s.n+'</option>';}});
h+='</select><button class="btn" onclick="addCcu();">+ CCU</button></div>';

if(PCcus.length>0){{h+='<div style="background:'+(ok?"rgba(105,240,174,0.05)":"rgba(255,82,82,0.05)")+';border:1px solid '+(ok?"rgba(105,240,174,0.2)":"rgba(255,82,82,0.2)")+';border-radius:8px;padding:14px;text-align:center;margin-top:16px;"><div style="font-size:13px;font-weight:700;color:'+(ok?"#69f0ae":"#ff5252")+';font-family:DM Mono,monospace;">'+(ok?"\u2713 PLAN AFFORDABLE \u2014 $"+rem+" to spare":"\u2717 $"+Math.abs(rem)+" SHORT")+'</div></div>';}}

const kept=fleet.filter((_,i)=>!PMelts.has(i)),targets=PCcus.map(c=>c.to);
if(kept.length+targets.length>0){{h+='<div class="sec-head" style="margin-top:16px;">FINAL FLEET</div>';
kept.forEach(s=>{{h+='<div class="card" style="border-left:2px solid #69f0ae;"><span class="tag" style="color:#69f0ae;background:rgba(105,240,174,0.1);border:1px solid rgba(105,240,174,0.2);">KEPT</span><span style="font-size:13px;font-weight:600;color:#eef1f5;">'+s.name+'</span></div>';}});
targets.forEach(t=>{{h+='<div class="card" style="border-left:2px solid #4fc3f7;"><span class="tag" style="color:#4fc3f7;background:rgba(79,195,247,0.1);border:1px solid rgba(79,195,247,0.2);">CCU</span><span style="font-size:13px;font-weight:600;color:#eef1f5;">'+t+'</span></div>';}});}}

area.innerHTML=h;}}
</script></body></html>'''

    return h


def build_overview(org, members, ships_db, images):
    """Generate fleet-overview.html — a visual fleet display with 3 themes."""
    
    # Collect all ships with metadata
    all_ships = []
    for member_name, ships in members.items():
        for entry in ships:
            sid = find_ship_id(entry["ship_name"], ships_db)
            ship_data = ships_db.get(sid, {}) if sid else {}
            size = ship_data.get("size", entry.get("size", "Medium"))
            role = ship_data.get("role", "Multi-Role")
            mfr = ship_data.get("mfr", entry.get("mfr", ""))
            name = ship_data.get("name", entry["ship_name"])
            img_b64 = images.get(sid, "") if sid else ""
            
            all_ships.append({
                "name": name,
                "mfr": mfr,
                "size": size,
                "role": role,
                "owner": member_name,
                "img": img_b64,
                "sid": sid or "",
            })
    
    # Sort: Capital > Large > Medium > Small, then by name
    size_order = {"Capital": 0, "Large": 1, "Medium": 2, "Small": 3, "Vehicle": 4}
    all_ships.sort(key=lambda s: (size_order.get(s["size"], 3), s["name"]))
    
    # Group by size
    tiers = {"Capital": [], "Large": [], "Medium": [], "Small": []}
    for s in all_ships:
        tier = s["size"] if s["size"] in tiers else "Medium"
        tiers[tier].append(s)
    
    total = len(all_ships)
    total_members = len(members)
    combat = sum(1 for s in all_ships if s["role"] in ("Combat", "Fighter", "Heavy Fighter", "Light Fighter", "Gunship", "Escort", "Bomber", "Patrol Gunship", "Multi-Role Fighter"))
    
    # Count manufacturers
    mfr_counts = {}
    for s in all_ships:
        m = s["mfr"] or "Unknown"
        mfr_counts[m] = mfr_counts.get(m, 0) + 1
    
    size_counts = {}
    for s in all_ships:
        size_counts[s["size"]] = size_counts.get(s["size"], 0) + 1
    
    org_name = org.get("name", "Fleet")
    org_motto = org.get("motto", "")
    
    # Build ship cards for holotable
    def holo_card(s):
        cls = {"Capital": "cap", "Large": "lrg", "Small": "sml"}.get(s["size"], "med")
        img_html = ""
        if s["img"]:
            img_html = f'<img class="ship-bg" src="{s["img"]}" alt="{s["name"]}">'
        return f"""<div class="holo-ship {cls}">
          {img_html}
          <div class="hs-content">
            <div class="hs-size">{s["size"].upper()}</div>
            <div class="hs-name">{s["name"]}</div>
            <div class="hs-mfr">{s["mfr"]}</div>
            <div class="hs-role">{s["role"]}</div>
            <div class="hs-owner">{s["owner"]}</div>
          </div>
        </div>"""
    
    def holo_tier(tier_name, ships):
        if not ships:
            return ""
        cards = "\n".join(holo_card(s) for s in ships)
        return f'<div class="holo-tier">{cards}</div>'
    
    holo_formation = "\n".join(holo_tier(t, tiers[t]) for t in ["Capital", "Large", "Medium", "Small"] if tiers[t])
    
    # CIC terminal rows
    cic_rows = ""
    for i, s in enumerate(all_ships):
        sz = {"Capital": "CAP", "Large": "LRG", "Small": "SML"}.get(s["size"], "MED")
        cursor = '<span class="cursor">_</span>' if i == len(all_ships) - 1 else ""
        cic_rows += f'<div class="cic-row"><div class="cic-c1">&gt; {s["name"].upper()}{cursor}</div><div class="cic-c2">{sz}</div><div class="cic-c3">{s["role"].upper()}</div><div class="cic-c4">{(s["mfr"] or "").upper()}</div></div>\n'
    
    # Tactical units
    tac_units = ""
    for s in all_ships:
        cls = {"Capital": "cap", "Large": "lrg", "Small": "sml"}.get(s["size"], "med")
        label = {"Capital": "CAP", "Large": "LRG", "Small": "SM"}.get(s["size"], "MED")
        img_html = ""
        if s["img"]:
            img_html = f'<img src="{s["img"]}" class="tac-thumb" alt="{s["name"]}">'
        tac_units += f"""<div class="tac-unit">
          <div class="tac-marker {cls}">{label}</div>
          {img_html}
          <div class="tac-info"><div class="tac-name">{s["name"]}</div>
          <div class="tac-detail">{s["mfr"]} // {s["role"]} // {s["size"]}</div>
          <div class="tac-owner">{s["owner"]}</div></div>
        </div>\n"""
    
    # Tactical sidebar
    tac_size_rows = "".join(f'<div class="tac-side-row"><div class="tac-side-k">{k}</div><div class="tac-side-v">{v}</div></div>' for k, v in size_counts.items())
    tac_mfr_rows = "".join(f'<div class="tac-side-row"><div class="tac-side-k">{k}</div><div class="tac-side-v">{v}</div></div>' for k, v in sorted(mfr_counts.items()))
    combat_level = "HIGH" if combat / max(total, 1) >= 0.5 else "MODERATE"
    
    h = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{org_name} — Fleet Overview</title>
<link href="https://fonts.googleapis.com/css2?family=Orbitron:wght@400;700;900&family=Rajdhani:wght@400;500;600;700&family=Share+Tech+Mono&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
:root{{
  --bg:#060a14;--bg2:#0d1020;--surface:rgba(14,18,36,0.9);
  --text:#e0e6f0;--text2:#8892a8;--text3:#4a5268;
  --accent:#00e0ff;--accent2:#0090aa;--accent-glow:rgba(0,224,255,0.15);
  --gold:#c89564;--green:#4ade80;--red:#f87171;--yellow:#fbbf24;
  --border:rgba(0,224,255,0.12);--border2:rgba(0,224,255,0.06);
}}
body{{background:var(--bg);color:var(--text);font-family:'Rajdhani',sans-serif;min-height:100vh;overflow-x:hidden;}}
body::before{{content:"";position:fixed;inset:0;pointer-events:none;background:linear-gradient(var(--border2) 1px,transparent 1px),linear-gradient(90deg,var(--border2) 1px,transparent 1px);background-size:40px 40px;opacity:0.5;z-index:0;}}
body::after{{content:"";position:fixed;inset:0;pointer-events:none;background:repeating-linear-gradient(0deg,transparent,transparent 2px,rgba(0,0,0,0.03) 2px,rgba(0,0,0,0.03) 4px);z-index:1;}}
.fleet-wrap{{position:relative;z-index:2;max-width:1100px;margin:0 auto;padding:20px;}}
.themes{{display:flex;gap:6px;margin-bottom:20px;justify-content:center;}}
.themes button{{font-family:'Share Tech Mono',monospace;font-size:12px;letter-spacing:2px;padding:8px 20px;border:1px solid var(--border);border-radius:4px;background:transparent;color:var(--text);cursor:pointer;text-transform:uppercase;transition:all 0.3s;}}
.themes button.active{{color:var(--accent);border-color:var(--accent);background:var(--accent-glow);}}
.back-link{{text-align:center;margin-bottom:16px;}}
.back-link a{{font-family:'Share Tech Mono',monospace;font-size:12px;color:var(--text2);text-decoration:none;letter-spacing:2px;}}
.back-link a:hover{{color:var(--accent);}}
.view{{display:none;}}.view.active{{display:block;}}

/* HOLOTABLE */
.holo-title{{font-family:'Orbitron';font-size:32px;font-weight:700;color:#00e0ff;letter-spacing:8px;text-align:center;text-shadow:0 0 40px rgba(0,224,255,0.3);margin-bottom:4px;}}
.holo-sub{{font-family:'Share Tech Mono';font-size:12px;color:#0090aa;letter-spacing:4px;text-align:center;margin-bottom:24px;}}
.holo-stats{{display:flex;justify-content:center;gap:40px;margin-bottom:28px;}}
.holo-stat-v{{font-family:'Orbitron';font-size:28px;font-weight:700;color:#00e0ff;text-align:center;}}
.holo-stat-l{{font-family:'Share Tech Mono';font-size:10px;color:#0090aa;letter-spacing:3px;text-align:center;}}
.holo-formation{{display:flex;flex-direction:column;align-items:center;gap:16px;}}
.holo-tier{{display:flex;gap:16px;justify-content:center;flex-wrap:wrap;}}
.holo-ship{{position:relative;background:rgba(0,224,255,0.03);border:1px solid rgba(0,224,255,0.15);display:flex;align-items:center;justify-content:center;text-align:center;transition:all 0.3s;overflow:hidden;}}
.holo-ship:hover{{background:rgba(0,224,255,0.08);border-color:rgba(0,224,255,0.5);box-shadow:0 0 30px rgba(0,224,255,0.15);}}
.holo-ship::before,.holo-ship::after{{content:"";position:absolute;width:12px;height:12px;border-color:#00e0ff;border-style:solid;border-width:0;opacity:0.5;}}
.holo-ship::before{{top:-1px;left:-1px;border-top-width:2px;border-left-width:2px;}}
.holo-ship::after{{bottom:-1px;right:-1px;border-bottom-width:2px;border-right-width:2px;}}
.holo-ship.cap{{width:300px;height:180px;border-radius:8px;}}
.holo-ship.lrg{{width:220px;height:140px;border-radius:6px;}}
.holo-ship.med{{width:170px;height:110px;border-radius:5px;}}
.holo-ship.sml{{width:140px;height:90px;border-radius:4px;}}
.ship-bg{{position:absolute;inset:0;width:100%;height:100%;object-fit:cover;opacity:0.2;z-index:0;}}
.holo-ship:hover .ship-bg{{opacity:0.35;}}
.hs-content{{position:relative;z-index:1;}}
.hs-name{{font-family:'Orbitron';font-weight:700;color:#e0e6f0;line-height:1.2;}}
.cap .hs-name{{font-size:17px;}}.lrg .hs-name{{font-size:14px;}}.med .hs-name{{font-size:12px;}}.sml .hs-name{{font-size:11px;}}
.hs-mfr{{font-family:'Share Tech Mono';color:#0090aa;font-size:10px;margin-top:3px;}}
.hs-role{{font-family:'Share Tech Mono';font-size:9px;color:#0090aa;letter-spacing:2px;text-transform:uppercase;margin-top:4px;}}
.hs-owner{{font-family:'Share Tech Mono';font-size:8px;color:rgba(0,224,255,0.3);margin-top:4px;}}
.hs-size{{position:absolute;top:6px;right:8px;font-family:'Share Tech Mono';font-size:7px;color:rgba(0,224,255,0.3);letter-spacing:1px;z-index:1;}}
@keyframes pulse{{0%,100%{{box-shadow:0 0 0 0 rgba(0,224,255,0.2);}}50%{{box-shadow:0 0 0 8px rgba(0,224,255,0);}}}}
.holo-ship.cap{{animation:pulse 3s ease-in-out infinite;}}

/* CIC */
.t-cic{{--bg:#0a0d00;--accent:#ffcc00;--accent2:#aa8800;--border:rgba(255,204,0,0.15);--text:#ffcc00;--text2:#aa8800;--text3:#665500;}}
.t-cic .cic-classified{{position:fixed;top:45%;left:50%;transform:translate(-50%,-50%) rotate(-30deg);font-family:'Orbitron';font-size:90px;font-weight:900;color:rgba(255,204,0,0.025);letter-spacing:20px;pointer-events:none;z-index:0;}}
.cic-header{{border:1px solid rgba(255,204,0,0.2);padding:14px 18px;margin-bottom:20px;position:relative;}}
.cic-header::before{{content:"TOP SECRET // FLEET MANIFEST";position:absolute;top:-8px;left:12px;background:var(--bg);padding:0 8px;font-size:10px;color:#aa8800;letter-spacing:3px;font-family:'Share Tech Mono';}}
.cic-owner{{font-size:20px;color:#ffcc00;letter-spacing:4px;font-family:'Share Tech Mono';}}
.cic-org{{font-size:11px;color:#886600;margin-top:4px;font-family:'Share Tech Mono';}}
.cic-terminal{{background:rgba(255,204,0,0.02);border:1px solid rgba(255,204,0,0.1);padding:16px;font-size:13px;line-height:2.2;font-family:'Share Tech Mono';position:relative;z-index:2;}}
.cic-row{{display:flex;border-bottom:1px solid rgba(255,204,0,0.06);padding:2px 0;}}
.cic-row:last-child{{border-bottom:none;}}
.cic-c1{{width:200px;color:#ffcc00;font-weight:bold;}}.cic-c2{{width:60px;color:#aa8800;text-align:center;}}.cic-c3{{flex:1;color:#886600;}}.cic-c4{{width:100px;color:#665500;text-align:right;}}
.cic-hdr{{color:#665500;font-size:10px;letter-spacing:2px;border-bottom:1px solid rgba(255,204,0,0.15);padding-bottom:4px;margin-bottom:4px;}}
.cic-summary{{display:flex;gap:20px;margin-top:20px;}}.cic-sbox{{border:1px solid rgba(255,204,0,0.15);padding:12px 18px;flex:1;text-align:center;}}
.cic-sv{{font-size:22px;color:#ffcc00;font-weight:bold;font-family:'Share Tech Mono';}}.cic-sl{{font-size:9px;color:#665500;letter-spacing:3px;margin-top:2px;font-family:'Share Tech Mono';}}
@keyframes blink{{0%,100%{{opacity:1;}}50%{{opacity:0;}}}}.cursor{{animation:blink 1s infinite;color:#ffcc00;}}

/* TACTICAL */
.t-tac{{--bg:#12141c;--accent:#60a5fa;--border:rgba(96,165,250,0.12);--text:#c8cdd8;--text2:#6b7280;}}
.t-tac::before{{background:linear-gradient(rgba(96,165,250,0.03) 1px,transparent 1px),linear-gradient(90deg,rgba(96,165,250,0.03) 1px,transparent 1px)!important;}}
.tac-header{{display:flex;justify-content:space-between;align-items:center;border-bottom:2px solid rgba(96,165,250,0.2);padding-bottom:14px;margin-bottom:20px;}}
.tac-title{{font-family:'Orbitron';font-size:22px;font-weight:700;color:#60a5fa;letter-spacing:4px;}}
.tac-badge{{font-family:'Share Tech Mono';font-size:10px;color:#3b82f6;letter-spacing:2px;border:1px solid rgba(96,165,250,0.3);padding:5px 14px;border-radius:3px;}}
.tac-grid{{display:grid;grid-template-columns:1fr 280px;gap:20px;position:relative;z-index:2;}}
@media(max-width:700px){{.tac-grid{{grid-template-columns:1fr;}}}}
.tac-formation{{display:flex;flex-direction:column;gap:8px;}}
.tac-unit{{display:flex;align-items:center;gap:12px;background:rgba(96,165,250,0.03);border:1px solid rgba(96,165,250,0.08);border-radius:6px;padding:12px 16px;transition:all 0.3s;}}
.tac-unit:hover{{border-color:rgba(96,165,250,0.3);background:rgba(96,165,250,0.06);}}
.tac-marker{{width:40px;height:40px;border:2px solid;display:flex;align-items:center;justify-content:center;font-family:'Share Tech Mono';font-size:10px;font-weight:bold;flex-shrink:0;}}
.tac-marker.cap{{border-color:#ef4444;color:#ef4444;border-radius:0;}}
.tac-marker.lrg{{border-color:#f59e0b;color:#f59e0b;border-radius:4px;}}
.tac-marker.med{{border-color:#60a5fa;color:#60a5fa;border-radius:50%;}}
.tac-marker.sml{{border-color:#6b7280;color:#6b7280;border-radius:50%;width:32px;height:32px;font-size:8px;}}
.tac-thumb{{width:40px;height:40px;object-fit:cover;border-radius:4px;}}
.tac-info{{flex:1;}}.tac-name{{font-size:16px;font-weight:700;color:#e0e4ea;}}.tac-detail{{font-family:'Share Tech Mono';font-size:10px;color:#6b7280;margin-top:2px;}}
.tac-owner{{font-family:'Share Tech Mono';font-size:9px;color:rgba(96,165,250,0.4);margin-top:2px;}}
.tac-sidebar{{background:rgba(96,165,250,0.02);border:1px solid rgba(96,165,250,0.08);border-radius:8px;padding:18px;}}
.tac-side-title{{font-family:'Share Tech Mono';font-size:10px;color:#6b7280;letter-spacing:3px;text-transform:uppercase;margin-bottom:12px;}}
.tac-side-row{{display:flex;justify-content:space-between;padding:7px 0;border-bottom:1px solid rgba(96,165,250,0.05);font-size:14px;}}
.tac-side-row:last-child{{border:none;}}.tac-side-k{{color:#6b7280;}}.tac-side-v{{color:#e0e4ea;font-weight:600;}}
.tac-threat{{margin-top:16px;padding:12px;border:1px solid rgba(34,197,94,0.2);border-radius:4px;background:rgba(34,197,94,0.03);}}
.tac-threat-title{{font-family:'Share Tech Mono';font-size:10px;color:#22c55e;letter-spacing:3px;margin-bottom:6px;}}
.tac-threat-val{{font-family:'Orbitron';font-size:26px;font-weight:700;color:#22c55e;}}
.tac-threat-sub{{font-family:'Share Tech Mono';font-size:10px;color:#6b7280;margin-top:2px;}}
.fleet-footer{{text-align:center;margin-top:30px;font-family:'Share Tech Mono';font-size:12px;color:var(--text2);letter-spacing:2px;line-height:2.2;}}
</style>
</head>
<body>
<div class="fleet-wrap">
  <div class="back-link"><a href="index.html">&larr; BACK TO FLEET LIST</a></div>
  <div class="themes">
    <button onclick="setTheme('holo')" class="active">Holotable</button>
    <button onclick="setTheme('cic')">CIC</button>
    <button onclick="setTheme('tac')">Tactical</button>
  </div>

  <div id="v-holo" class="view active">
    <div class="holo-title">{org_name}</div>
    <div class="holo-sub">{org_motto}</div>
    <div class="holo-stats">
      <div><div class="holo-stat-v">{total}</div><div class="holo-stat-l">Ships</div></div>
      <div><div class="holo-stat-v">{total_members}</div><div class="holo-stat-l">Members</div></div>
      <div><div class="holo-stat-v">{combat}</div><div class="holo-stat-l">Combat</div></div>
    </div>
    <div class="holo-formation">
      {holo_formation}
    </div>
  </div>

  <div id="v-cic" class="view">
    <div class="cic-classified">CLASSIFIED</div>
    <div class="cic-header">
      <div class="cic-owner">{org_name.upper()}</div>
      <div class="cic-org">{org_motto.upper() if org_motto else ""}</div>
    </div>
    <div class="cic-terminal">
      <div class="cic-hdr"><div class="cic-row"><div class="cic-c1">DESIGNATION</div><div class="cic-c2">CLASS</div><div class="cic-c3">ROLE</div><div class="cic-c4">MFR</div></div></div>
      {cic_rows}
    </div>
    <div class="cic-summary">
      <div class="cic-sbox"><div class="cic-sv">{total}</div><div class="cic-sl">TOTAL ASSETS</div></div>
      <div class="cic-sbox"><div class="cic-sv">{total_members}</div><div class="cic-sl">MEMBERS</div></div>
      <div class="cic-sbox"><div class="cic-sv">{round(combat/max(total,1)*100)}%</div><div class="cic-sl">COMBAT RATIO</div></div>
      <div class="cic-sbox"><div class="cic-sv">OPERATIONAL</div><div class="cic-sl">FLEET STATUS</div></div>
    </div>
  </div>

  <div id="v-tac" class="view">
    <div class="tac-header">
      <div>
        <div class="tac-title">Force Composition</div>
        <div style="font-family:'Share Tech Mono';font-size:11px;color:#6b7280;margin-top:4px;">{org_name.upper()}{(" // " + org_motto.upper()) if org_motto else ""}</div>
      </div>
      <div class="tac-badge">FLEET STATUS: OPERATIONAL</div>
    </div>
    <div class="tac-grid">
      <div class="tac-formation">{tac_units}</div>
      <div class="tac-sidebar">
        <div class="tac-side-title">Fleet summary</div>
        <div class="tac-side-row"><div class="tac-side-k">Total ships</div><div class="tac-side-v">{total}</div></div>
        <div class="tac-side-row"><div class="tac-side-k">Members</div><div class="tac-side-v">{total_members}</div></div>
        {tac_size_rows}
        <div class="tac-side-title" style="margin-top:16px;">Manufacturers</div>
        {tac_mfr_rows}
        <div class="tac-threat">
          <div class="tac-threat-title">Combat readiness</div>
          <div class="tac-threat-val">{combat_level}</div>
          <div class="tac-threat-sub">{combat} of {total} combat-capable</div>
        </div>
      </div>
    </div>
  </div>

  <div class="fleet-footer">{org_name} &mdash; {org_motto}<br>BUILD {BUILD_VERSION}<br><a href="https://buymeacoffee.com/schindi21" target="_blank" style="color:var(--text2);text-decoration:none;border-bottom:1px solid var(--text3);">buy me a coffee</a> &bull; BTC: <span style="font-size:10px;color:var(--text2);">bc1qdyh5g2zska7s9e4vu27hzqyyre60t6khl4srnx</span></div>
</div>
<script>
function setTheme(t){{
  document.body.className=t==="cic"?"t-cic":t==="tac"?"t-tac":"";
  document.querySelectorAll(".view").forEach(function(v){{v.classList.remove("active");}});
  document.getElementById("v-"+t).classList.add("active");
  document.querySelectorAll(".themes button").forEach(function(b){{b.classList.remove("active");}});
  event.target.classList.add("active");
}}
</script>
</body></html>"""
    
    return h


def open_enhancement():
    """Return the enhancement script block (theme toggle, readability, ship images)."""
    return """<script>
(function(){
  var css=document.createElement("style");
  css.textContent=`
    body.light{background:#f5f0e8!important;color:#1a1a2e!important;}
    body.light::before{opacity:0.02!important;background-image:linear-gradient(rgba(139,105,20,0.15) 1px,transparent 1px),linear-gradient(90deg,rgba(139,105,20,0.15) 1px,transparent 1px)!important;}
    body.light .glow{background:radial-gradient(ellipse,rgba(139,105,20,0.04) 0%,transparent 70%)!important;}
    body.light .org-name{color:#1a1a2e!important;}
    body.light .card{background:rgba(255,255,255,0.9)!important;border-color:rgba(0,0,0,0.08)!important;}
    body.light .ship-name{color:#1a1a2e!important;}
    body.light .ship-owner{color:#777!important;}body.light .ship-owner span{color:#555!important;}
    body.light .detail-key{color:#888!important;}body.light .detail-val{color:#555!important;}
    body.light .cat-header{color:#8b6914!important;border-color:rgba(0,0,0,0.08)!important;}
    body.light .tab{background:rgba(0,0,0,0.02)!important;border-color:rgba(0,0,0,0.1)!important;color:#888!important;}
    body.light .tab.active{background:rgba(139,105,20,0.08)!important;border-color:rgba(139,105,20,0.3)!important;color:#8b6914!important;}
    body.light .sort-btn,body.light .mbr-btn{background:rgba(0,0,0,0.02)!important;border-color:rgba(0,0,0,0.1)!important;color:#888!important;}
    body.light .sort-btn.active,body.light .mbr-btn.active{background:rgba(139,105,20,0.08)!important;border-color:rgba(139,105,20,0.3)!important;color:#8b6914!important;}
    body.light .summary{background:rgba(139,105,20,0.03)!important;border-color:rgba(139,105,20,0.1)!important;}
    body.light .hl{color:#555!important;background:rgba(0,0,0,0.03)!important;border-color:rgba(0,0,0,0.08)!important;}
    body.light .details-toggle{border-color:rgba(0,0,0,0.06)!important;}
    body.light .member-name{color:#1a1a2e!important;}
    body.light .hero-fade{background:linear-gradient(180deg,transparent 30%,rgba(255,255,255,0.95) 100%)!important;}
    body.light .footer{color:#999!important;}
    .org-motto{font-size:11px!important;opacity:0.9!important;}
    .org-desc{font-size:12px!important;color:#8899aa!important;}
    .s-val{font-size:20px!important;}.s-lbl{font-size:9px!important;color:#667!important;}
    .tab{font-size:11px!important;padding:8px 14px!important;}
    .sort-btn,.mbr-btn{font-size:10px!important;padding:6px 12px!important;}
    .sort-label{font-size:10px!important;}.cat-header{font-size:13px!important;}
    .mfr-tag{font-size:10px!important;}.role-tag{font-size:9px!important;color:#778!important;}
    .size-tag{font-size:9px!important;}.ship-name{font-size:20px!important;}
    .ship-owner{font-size:11px!important;color:#778!important;}.ship-owner span{color:#99a!important;}
    .ship-desc{font-size:12px!important;}.hl{font-size:10px!important;padding:4px 10px!important;}
    .details-toggle{font-size:11px!important;}.detail-key{font-size:9px!important;color:#667!important;}
    .detail-val{font-size:12px!important;color:#99a!important;}.member-count{font-size:11px!important;}
    .footer{font-size:12px!important;opacity:0.7!important;line-height:2.2!important;}
    #sc-theme-toggle{position:fixed;top:12px;right:12px;z-index:100;padding:6px 14px;font-size:10px;letter-spacing:2px;font-family:"DM Mono",monospace;border:1px solid rgba(200,149,100,0.3);border-radius:4px;background:rgba(200,149,100,0.08);color:#c89564;cursor:pointer;}
    body.light #sc-theme-toggle{border-color:rgba(139,105,20,0.3);background:rgba(139,105,20,0.06);color:#8b6914;}
    .wiki-hero{position:relative;width:100%;height:140px;overflow:hidden;}
    .wiki-hero img{width:100%;height:100%;object-fit:cover;display:block;}
    .wiki-hero .hero-fade{position:absolute;inset:0;background:linear-gradient(180deg,transparent 30%,rgba(8,12,20,0.95) 100%);}
    body.light .wiki-hero .hero-fade{background:linear-gradient(180deg,transparent 30%,rgba(255,255,255,0.95) 100%);}
  `;
  document.head.appendChild(css);
  var btn=document.createElement("div");btn.id="sc-theme-toggle";
  var isLight=localStorage.getItem("sc-viewer-theme")==="light";
  if(isLight)document.body.classList.add("light");
  function updateLabel(){btn.textContent=isLight?"\u263d DARK":"\u263c LIGHT";}
  updateLabel();
  btn.addEventListener("click",function(){isLight=!isLight;document.body.classList.toggle("light");localStorage.setItem("sc-viewer-theme",isLight?"light":"dark");updateLabel();});
  document.body.appendChild(btn);
  var cards=document.querySelectorAll(".card");
  var shipNames=[];
  cards.forEach(function(c){var nm=c.querySelector(".ship-name");if(nm)shipNames.push({el:c,name:nm.textContent.trim()});});
  if(shipNames.length===0)return;
  var mfrs=["RSI","Aegis","Anvil","Argo","Aopoa","Banu","Consolidated","Crusader","Drake","Esperia","Gatac","Kruger","MISC","Origin","Tumbril"];
  function stripMfr(n){var parts=n.split(" ");if(parts.length>1&&mfrs.some(function(m){return parts[0].toLowerCase()===m.toLowerCase();}))return parts.slice(1).join(" ");return n;}
  var wikiMap={};
  shipNames.forEach(function(s){wikiMap[stripMfr(s.name)]=s;});
  var titles=Object.keys(wikiMap).join("|");
  console.log("[Fleet Viewer] Fetching images for: "+titles);
  fetch("https://starcitizen.tools/api.php?action=query&titles="+encodeURIComponent(titles)+"&prop=pageimages&format=json&pithumbsize=600&origin=*")
    .then(function(r){return r.json();})
    .then(function(data){
      var pages=data.query&&data.query.pages||{};
      var found=0;
      Object.keys(pages).forEach(function(pid){
        var page=pages[pid];
        if(page.thumbnail&&page.thumbnail.source){
          var entry=wikiMap[page.title];if(!entry)return;
          var noImg=entry.el.querySelector(".no-img");if(!noImg)return;
          var hero=document.createElement("div");hero.className="wiki-hero";
          var img=document.createElement("img");img.src=page.thumbnail.source;img.alt=entry.name;img.loading="lazy";
          img.onerror=function(){hero.style.display="none";};
          var fade=document.createElement("div");fade.className="hero-fade";
          hero.appendChild(img);hero.appendChild(fade);
          noImg.replaceWith(hero);
          var info=entry.el.querySelector(".info");if(info)info.classList.add("has-img");
          found++;
        }
      });
      console.log("[Fleet Viewer] Loaded "+found+" ship images");
    }).catch(function(e){console.log("[Fleet Viewer] Wiki image fetch failed:",e);});
})();
""" + "<" + "/script>"


def download_ship_images(members, ships_db):
    """Download ship images from starcitizen.tools wiki during build."""
    import urllib.request, urllib.parse
    
    images_dir = "images"
    os.makedirs(images_dir, exist_ok=True)
    
    # Strip manufacturer prefixes for wiki lookup
    mfrs = ["RSI", "Aegis", "Anvil", "Argo", "Aopoa", "Banu", "Consolidated Outland",
            "Crusader", "Drake", "Esperia", "Gatac", "Kruger", "MISC", "Origin", "Tumbril"]
    
    def strip_mfr(name):
        for m in mfrs:
            if name.startswith(m + " "):
                return name[len(m)+1:]
        return name
    
    # Collect all unique ship names + resolved names from ships.json
    all_ships = set()
    ship_name_variants = {}
    for member_name, ships in members.items():
        for entry in ships:
            name = entry["ship_name"]
            all_ships.add(name)
            variants = {name}
            sid = _resolve_ship_id(name, ships_db)
            if sid and sid in ships_db:
                full_name = ships_db[sid]["name"]
                variants.add(full_name)
                variants.add(strip_mfr(full_name))
                # Also find related ships (e.g., 600i -> 600i Explorer, 600i Touring)
                for other_sid, other_data in ships_db.items():
                    if other_sid.startswith(sid + "_") or other_sid == sid:
                        variants.add(other_data["name"])
                        variants.add(strip_mfr(other_data["name"]))
            variants.add(strip_mfr(name))
            ship_name_variants[name] = variants
    
    if not all_ships:
        return {}
    
    wiki_map = {}
    for name in all_ships:
        variants = ship_name_variants.get(name, {name, strip_mfr(name)})
        for v in variants:
            if v not in wiki_map:
                wiki_map[v] = name
    print(f"  Wiki titles to query: {list(wiki_map.keys())}")
    
    # Check which images we already have
    existing = {}
    for name in all_ships:
        sid = _resolve_ship_id(name, ships_db) or name.replace(" ", "_").replace("/", "_").lower()
        for ext in [".jpg", ".png", ".webp"]:
            path = os.path.join(images_dir, sid + ext)
            if os.path.exists(path):
                existing[name] = path
                break
    
    if len(existing) == len(all_ships):
        print(f"  All {len(existing)} ship images already cached")
        return existing
    
    # Fetch from wiki API
    titles = "|".join(wiki_map.keys())
    api_url = f"https://starcitizen.tools/api.php?action=query&titles={urllib.parse.quote(titles)}&prop=pageimages&format=json&pithumbsize=600&origin=*"
    
    print(f"  Fetching ship images from starcitizen.tools...")
    try:
        req = urllib.request.Request(api_url, headers={"User-Agent": "SC-Fleet-Viewer/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            import json as _json
            data = _json.loads(resp.read().decode())
        
        pages = data.get("query", {}).get("pages", {})
        downloaded = 0
        for pid, page in pages.items():
            thumb = page.get("thumbnail", {}).get("source")
            title = page.get("title", "")
            if not thumb or title not in wiki_map:
                continue
            
            ship_name = wiki_map[title]
            if ship_name in existing:
                continue
            
            # Download the image
            # Use ship_id for filename so collect_images finds it
            sid = _resolve_ship_id(ship_name, ships_db) or ship_name.replace(" ", "_").replace("/", "_").lower()
            ext = ".jpg"
            if ".png" in thumb.lower():
                ext = ".png"
            elif ".webp" in thumb.lower():
                ext = ".webp"
            
            img_path = os.path.join(images_dir, sid + ext)
            try:
                req2 = urllib.request.Request(thumb, headers={"User-Agent": "SC-Fleet-Viewer/1.0"})
                with urllib.request.urlopen(req2, timeout=10) as img_resp:
                    with open(img_path, "wb") as img_f:
                        img_f.write(img_resp.read())
                existing[ship_name] = img_path
                downloaded += 1
                print(f"    ✓ {ship_name}")
            except Exception as e:
                print(f"    ✗ {ship_name}: {e}")
        
        print(f"  Downloaded {downloaded} new images ({len(existing)} total cached)")
    except Exception as e:
        print(f"  Wiki API request failed: {e}")
    
    # Try individual lookups for ships not found in batch
    for short_name, ship_name in wiki_map.items():
        if ship_name in existing:
            continue
        try:
            url2 = f"https://starcitizen.tools/api.php?action=query&titles={urllib.parse.quote(short_name)}&prop=pageimages&format=json&pithumbsize=600&origin=*"
            req = urllib.request.Request(url2, headers={"User-Agent": "SC-Fleet-Viewer/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = _json.loads(resp.read().decode())
            pages = data.get("query", {}).get("pages", {})
            for pid, page in pages.items():
                thumb = page.get("thumbnail", {}).get("source")
                if thumb:
                    sid = _resolve_ship_id(ship_name, ships_db) or ship_name.replace(" ", "_").replace("/", "_").lower()
                    ext = ".jpg" if ".jpg" in thumb.lower() else ".png"
                    img_path = os.path.join(images_dir, sid + ext)
                    req2 = urllib.request.Request(thumb, headers={"User-Agent": "SC-Fleet-Viewer/1.0"})
                    with urllib.request.urlopen(req2, timeout=10) as img_resp:
                        with open(img_path, "wb") as img_f:
                            img_f.write(img_resp.read())
                    existing[ship_name] = img_path
                    print(f"    ✓ {ship_name} (individual)")
        except:
            pass
    
    return existing


def main():
    print("SC Fleet Viewer — Build")
    print(f"Version: {BUILD_VERSION}")
    print("=" * 40)

    print(f"Loading ship database ({SHIPS_FILE})...")
    ships_db = load_ships_db()
    print(f"  {len(ships_db)} ships in database")

    # Try org.json first (simpler), fall back to Excel
    org_json = load_org_json()
    
    # Try CSV imports from hangars/ folder first
    print(f"Checking for hangar CSVs (hangars/)...")
    csv_members = load_hangars_csv(ships_db)
    
    if csv_members:
        members = csv_members
        org = org_json or {"name": "My Org", "motto": "", "description": "", "logo": ""}
        print(f"  Using CSV import mode")
    elif os.path.exists(EXCEL_FILE):
        print(f"  No CSVs found, falling back to Excel ({EXCEL_FILE})...")
        org_excel, members = load_excel()
        org = org_json or org_excel
    else:
        print(f"  No fleet data found!")
        print(f"  Place CSV files in hangars/ (exported from SC Fleet Planner → Export for Org)")
        sys.exit(1)

    total_ships = sum(len(ships) for ships in members.values())
    print(f"  Org: {org['name']}")
    print(f"  Members: {len(members)}")
    print(f"  Total ships: {total_ships}")

    print(f"Loading images ({IMAGES_DIR}/)...")
    os.makedirs(IMAGES_DIR, exist_ok=True)
    # Download ship images from wiki if not cached
    wiki_images = download_ship_images(members, ships_db)
    images = collect_images(members, ships_db)
    # Merge wiki images (collect_images handles local files, wiki_images adds wiki downloads)
    print(f"  {len(images)} local images, {len(wiki_images)} wiki images")

    print("Generating HTML...")
    html = build_html(org, members, ships_db, images)

    # Post-process: escape </ in script blocks to prevent HTML parser breakage
    import re as _re
    def _fix_slash(m):
        return "<script>" + m.group(1).replace("</", "<\\/") + "</script>"
    html = _re.sub(r'<script>(.*?)</script>', _fix_slash, html, flags=_re.DOTALL)

    # Inject enhancement (theme toggle, readability, ship images)
    enh = open_enhancement()
    html = html.replace("</body></html>", enh + "\n</body></html>")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(html)

    size_kb = len(html) // 1024
    print(f"\n✓ Generated {OUTPUT_FILE} ({size_kb} KB)")
    print(f"  Open index.html to view your org fleet!")

    # Generate fleet overview (visual display)
    print("Generating fleet-overview.html...")
    overview = build_overview(org, members, ships_db, images)
    with open("fleet-overview.html", "w", encoding="utf-8") as f:
        f.write(overview)
    print(f"\u2713 Generated fleet-overview.html ({len(overview) // 1024} KB)")


if __name__ == "__main__":
    main()
